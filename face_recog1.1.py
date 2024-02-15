import cv2
import time
import openpyxl
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Define some constants
ATTENDANCE_INTERVAL = 30  # in seconds
VIDEO_CODEC = cv2.VideoWriter_fourcc(*'XVID')
VIDEO_FPS = 10
VIDEO_SIZE = (640, 480)

# Create some variables
attendance_recorded = False
attendance_time = time.time()

# Create a video capture object
cap = cv2.VideoCapture(0)
cap.set(cv2.CAP_PROP_FPS, VIDEO_FPS)
cap.set(cv2.CAP_PROP_FRAME_WIDTH, VIDEO_SIZE[0])
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, VIDEO_SIZE[1])

# Get the video dimension and create the video writer
width, height = VIDEO_SIZE
out = cv2.VideoWriter("output_video_file\_video.avi", VIDEO_CODEC, VIDEO_FPS, (width, height))

# Load the Haar cascade file and the font
face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
font = cv2.FONT_HERSHEY_SIMPLEX

# Check if the camera opened successfully
if not cap.isOpened():
    print("unable to read camera feed")
    exit()

# Initialize the face recognition model
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(r'Train_model\face_recognizer_model.yml')

# Create an Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Set up the header row in the Excel worksheet
worksheet.append(['Name', 'Date', 'Time', 'Status', 'Email'])

# Load the names and email addresses from the model
names = []
emails = []
with open('student_data\data.txt', 'r') as f:
    for line in f:
        name, email = line.strip().split(',')
        names.append(name)
        emails.append(email)
        print(emails)

        
# Create a list of all students
all_students = names.copy()


# Email configuration
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USERNAME = 'piyanshul1307@gmail.com'
SMTP_PASSWORD = 'gofwnjwjzammznsf'

# Function to send email
def send_email(to_email, subject, body):
    from_email = SMTP_USERNAME

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    try:
        
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            print("Connected to SMTP server.")
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            print("Logged in to SMTP server.")
            server.sendmail(from_email, to_email, msg.as_string())
            print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"Error sending email to {to_email}: {e}")

    # Using an infinite loop to read the frame, detect faces, and write time to the feed
while True:
    ret, frame = cap.read()
    if not ret:
        print("unable to read camera feed")
        break

    # Convert the frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Detect faces on the frame
    faces = face_cascade.detectMultiScale(gray, 1.1, 4)

    # Loop through the detected faces
    for (x, y, w, h) in faces:
        # Draw a rectangle around the face
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        # Recognize the face
        label, confidence = recognizer.predict(cv2.resize(gray[y:y + h, x:x + w], (100, 100)))

        # Write the name, date, and time to the Excel worksheet
        if not attendance_recorded:
            current_time_string = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            attendance_time = time.mktime(time.strptime(current_time_string, "%Y-%m-%d %H:%M:%S"))
            status = "Present"

            # Check if the student is absent
            if names[label] in worksheet.cell(row=1, column=1).value and worksheet.cell(row=1, column=1).value != 'Name':
                status = "Absent"
                # Send an email to absent students
                send_email(emails[label], "Attendance Alert", f"Dear {names[label]},\n\nYou were absent in today's class. Please contact us for more details.")

            worksheet.append([names[label], time.strftime("%Y-%m-%d"), time.strftime("%H:%M:%S"), status, emails[label]])
            attendance_recorded = True

            # Remove the recognized student from the list
            removed_student = all_students.pop(all_students.index(names[label]))
            print(f"PRESENT: {removed_student}")

        # Display the name on the frame
        cv2.putText(frame, names[label], (x, y - 5), font, 0.5, (255, 0, 0), 1, cv2.LINE_AA)

    # Write the frame in the video
    out.write(frame)

    # Display the resulting frame
    cv2.imshow("video footage", frame)

    # Check if the user pressed 'q'
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

    # Check if it's time to record attendance again
    if attendance_recorded and time.time() - attendance_time > ATTENDANCE_INTERVAL:
        attendance_recorded = False

# Add rows for unrecognized faces with status "Absent"
for i in range(2, worksheet.max_row + 1):
    if worksheet.cell(row=i, column=3).value is None:
        worksheet.cell(row=i, column=4).value = "Absent"



# Add rows for unrecognized faces with status "Absent"
for i, student in enumerate(all_students, start=2):
    worksheet.append([student, '', '', "Absent"])
    print(f"ABSENT : {student}")


# Save the Excel workbook
workbook.save(r'attendence_file\attendance.xlsx')

# Close all the windows
cv2.destroyAllWindows()

# Close the video writer
out.release()

# Close the video capture
cap.release()
